import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

input_file = "team_kpis_mock_1.xlsx"
df = pd.read_excel(input_file)

# Passo 1: salvar dados brutos com validação e formatação na aba "All Data"
temp_file = "kpi_temp.xlsx"
df.to_excel(temp_file, index=False)

wb = load_workbook(temp_file)
ws = wb.active
ws.title = "All Data"

headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
status_col = get_column_letter(headers["Status"])
responsible_col = get_column_letter(headers["Responsible"])
category_col = get_column_letter(headers["Category"])
result_col = get_column_letter(headers["Current Result"])
target_col = get_column_letter(headers["Target"])
deadline_col = get_column_letter(headers["Deadline"])

row_start = 2
row_end = ws.max_row

responsibles = ["John", "Michael", "Sarah", "Jessica", "David", "Emily", "James", "Emma", "Robert", "Olivia"]
statuses = ["Achieved", "In Progress", "At Risk", "Delayed"]
categories = ["Sales", "Marketing", "Support", "Finance", "Development", "HR", "Logistics"]

dv_responsible = DataValidation(type="list", formula1=f'"{",".join(responsibles)}"', allow_blank=True)
dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
dv_category = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)

ws.add_data_validation(dv_responsible)
ws.add_data_validation(dv_status)
ws.add_data_validation(dv_category)

dv_responsible.add(f"{responsible_col}{row_start}:{responsible_col}{row_end}")
dv_status.add(f"{status_col}{row_start}:{status_col}{row_end}")
dv_category.add(f"{category_col}{row_start}:{category_col}{row_end}")

status_colors = {
    "Achieved": "C6EFCE",
    "In Progress": "FFEB9C",
    "At Risk": "FCE4D6",
    "Delayed": "F8CBAD"
}

for status in statuses:
    ws.conditional_formatting.add(
        f"{status_col}{row_start}:{status_col}{row_end}",
        FormulaRule(
            formula=[f'${status_col}{row_start}="{status}"'],
            fill=PatternFill(start_color=status_colors[status], end_color=status_colors[status], fill_type="solid")
        )
    )

ws.conditional_formatting.add(
    f"{result_col}{row_start}:{result_col}{row_end}",
    FormulaRule(
        formula=[f"${result_col}{row_start}<${target_col}{row_start}"],
        fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
    )
)

ws.conditional_formatting.add(
    f"{deadline_col}{row_start}:{deadline_col}{row_end}",
    FormulaRule(
        formula=[f'AND(${deadline_col}{row_start}-TODAY()<=7, ${deadline_col}{row_start}-TODAY()>0)'],
        fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    )
)

wb.save(temp_file)
wb.close()

# Passo 2: abrir com pandas ExcelWriter para adicionar abas filtradas (sem a coluna Status)
with pd.ExcelWriter(temp_file, mode="a", engine="openpyxl", if_sheet_exists="replace") as writer:
    for status in statuses:
        df_status = df[df["Status"] == status].copy()
        if df_status.empty:
            continue
        df_status = df_status.drop(columns=["Status"])
        df_status.sort_values(by="Deadline", inplace=True)
        df_status.to_excel(writer, sheet_name=status, index=False)

# Opcional: renomear o arquivo final
import os
final_file = "kpi_report_data_filtered_sheets.xlsx"
os.rename(temp_file, final_file)

print(f"✅ File '{final_file}' created with all requirements!")
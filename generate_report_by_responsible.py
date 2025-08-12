import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
import os

input_file = "mock_data.xlsx"
df = pd.read_excel(input_file)

responsibles = df["Responsible"].dropna().unique()
statuses = df["Status"].dropna().unique()
categories = df["Category"].dropna().unique()

responsibles = [str(r) for r in responsibles]
statuses = [str(s) for s in statuses]
categories = [str(c) for c in categories]

def apply_dropdown_and_formatting(ws, df_sheet, include_responsible=True):
    row_start = 2
    row_end = ws.max_row
    headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    if include_responsible and "Responsible" in headers:
        col_resp = get_column_letter(headers["Responsible"])
        dv_resp = DataValidation(type="list", formula1=f'"{",".join(responsibles)}"', allow_blank=True)
        ws.add_data_validation(dv_resp)
        dv_resp.add(f"{col_resp}{row_start}:{col_resp}{row_end}")

    if "Status" in headers:
        col_status = get_column_letter(headers["Status"])
        dv_status = DataValidation(type="list", formula1=f'"{",".join(statuses)}"', allow_blank=True)
        ws.add_data_validation(dv_status)
        dv_status.add(f"{col_status}{row_start}:{col_status}{row_end}")

    if "Category" in headers:
        col_cat = get_column_letter(headers["Category"])
        dv_cat = DataValidation(type="list", formula1=f'"{",".join(categories)}"', allow_blank=True)
        ws.add_data_validation(dv_cat)
        dv_cat.add(f"{col_cat}{row_start}:{col_cat}{row_end}")

    status_colors = {
        "Achieved": "C6EFCE",
        "In Progress": "FFEB9C",
        "At Risk": "FCE4D6",
        "Delayed": "F8CBAD"
    }
    if "Status" in headers:
        col_status = get_column_letter(headers["Status"])
        for status, color in status_colors.items():
            ws.conditional_formatting.add(
                f"{col_status}{row_start}:{col_status}{row_end}",
                FormulaRule(
                    formula=[f'${col_status}{row_start}="{status}"'],
                    fill=PatternFill(start_color=color, end_color=color, fill_type="solid")
                )
            )

    if "Current Result" in headers and "Target" in headers:
        col_result = get_column_letter(headers["Current Result"])
        col_target = get_column_letter(headers["Target"])
        ws.conditional_formatting.add(
            f"{col_result}{row_start}:{col_result}{row_end}",
            FormulaRule(
                formula=[f"${col_result}{row_start}<${col_target}{row_start}"],
                fill=PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            )
        )

    if "Deadline" in headers:
        col_deadline = get_column_letter(headers["Deadline"])
        ws.conditional_formatting.add(
            f"{col_deadline}{row_start}:{col_deadline}{row_end}",
            FormulaRule(
                formula=[f'AND(${col_deadline}{row_start}-TODAY()<=7, ${col_deadline}{row_start}-TODAY()>0)'],
                fill=PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            )
        )

# Salva todas as abas de uma vez com ExcelWriter
temp_file = "kpi_temp.xlsx"
with pd.ExcelWriter(temp_file, engine="openpyxl") as writer:
    # Aba All Data (com Responsible)
    df.to_excel(writer, sheet_name="All Data", index=False)

    # Abas individuais sem Responsible
    for resp in responsibles:
        df_resp = df[df["Responsible"] == resp].copy()
        if df_resp.empty:
            continue
        df_resp = df_resp.drop(columns=["Responsible"])
        df_resp.sort_values(by="Deadline", inplace=True)
        df_resp.to_excel(writer, sheet_name=str(resp), index=False)

# Agora abre para aplicar formatação
wb = load_workbook(temp_file)

# Formata "All Data" com Responsible
ws_all = wb["All Data"]
apply_dropdown_and_formatting(ws_all, df, include_responsible=True)

# Formata abas individuais sem Responsible
for resp in responsibles:
    if resp not in wb.sheetnames:
        continue
    ws_resp = wb[resp]
    df_resp = df[df["Responsible"] == resp].copy()
    df_resp = df_resp.drop(columns=["Responsible"])
    apply_dropdown_and_formatting(ws_resp, df_resp, include_responsible=False)

final_file = "kpi_reports.xlsx"
wb.save(final_file)
wb.close()
os.remove(temp_file)

print(f"✅ Final file saved as '{final_file}' with full formatting and dropdowns!")